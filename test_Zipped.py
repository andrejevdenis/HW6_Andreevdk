import os.path

from openpyxl import load_workbook
from zipfile import ZipFile
from PyPDF2 import PdfReader
import pytest

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
print(CURRENT_DIR)

TEMP_DIR = os.path.join(CURRENT_DIR, 'temp')
print(TEMP_DIR)

with ZipFile(fr'{TEMP_DIR}\metanit.zip', "w") as myzip:
    myzip.write('1.xlsx')
    myzip.write('2.csv')
    myzip.write('3.pdf')
    print(myzip.infolist())
    # Pdffile = PdfReader(fr'{CURRENT_DIR}\3.pdf')

    with myzip.open('1.xlsx') as Excell_file:
        content = Excell_file
        sheet = load_workbook(content).active
        print(sheet.cell(row=3, column=2).value)

    with myzip.open('3.pdf') as Pdf_file:
        content = PdfReader(Pdf_file)
        print(len(content.pages))

    with myzip.open('2.csv') as Csv_file:
        text = Csv_file.read().decode('UTF-8')
        print(text)

def test_csv():
    assert '105;2005' in text

def test_pdf():
    assert len(content.pages) == 13

def test_xlsx():
    assert sheet.cell(row=3, column=2).value == 'Март'
