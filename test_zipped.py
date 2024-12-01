import os.path
import shutil

import pytest
from openpyxl import load_workbook
from zipfile import ZipFile
from PyPDF2 import PdfReader

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
print(CURRENT_DIR)

TEMP_DIR = os.path.join(CURRENT_DIR, 'temp')
print(TEMP_DIR)

with ZipFile(fr'{TEMP_DIR}\metanit.zip', "w") as myzip:
    myzip.write('1.xlsx')
    myzip.write('2.csv')
    myzip.write('3.pdf')
    print(myzip.infolist())

    Excell_file = myzip.open('1.xlsx')
    content = Excell_file
    sheet = load_workbook(content).active

    Pdf_file = myzip.open('3.pdf')
    content1 = PdfReader(Pdf_file)

    Csv_file = myzip.open('2.csv')
    text = Csv_file.read().decode('UTF-8')

    def test_xlsx():
        assert sheet.cell(row=3, column=2).value == 'Март'
        Excell_file.close()
        content.close()

    def test_pdf():
        assert "zipfile.is_zipfile(filename)" in content1.pages[1].extract_text()
        Pdf_file.close()

    def test_csv():
        assert '105;2005' in text
        Csv_file.close()

# ZipFile(fr'{TEMP_DIR}\metanit.zip', 'w').close()

# shutil.rmtree(os.path.join(CURRENT_DIR, "temp"))
os.remove(fr'{TEMP_DIR}\metanit.zip')